#!/usr/bin/env python3
"""
Supplementary Tables 7 + 8: De novo mutation enrichment at gene
conversion-compatible positions, all-mappable (ST7) and outside segmental
duplications (ST8).

Reads ALL_StatsSummary.R.out.txt files produced by step 06 on the DNM input
(all DNMs, not parent-stratified), in stats_allmapp (ST7) and stats_nosegdupmapp
(ST8). Assembles them into the published one-panel xlsx layout.

Input:
  <PROJECT_ROOT>/geneconv_complete/dnm_analysis/export/stats_allmapp/
    _All_1000000_1000000_ALL_StatsSummary.R.out.txt          (ST7)
  <PROJECT_ROOT>/geneconv_complete/dnm_analysis/export/stats_nosegdupmapp/
    _All_1000000_1000000_ALL_StatsSummary.R.out.txt          (ST8)

Output:
  <PROJECT_ROOT>/geneconv_complete/__PAPER_3/TablesFinal/SupplementaryTable_{7,8}.xlsx
"""

import math
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Border, Side

PROJECT_ROOT = Path(os.environ.get("PROJECT_ROOT", ""))
if not PROJECT_ROOT.is_dir():
    sys.exit("PROJECT_ROOT env var must point to a directory")

DNM_BASE = PROJECT_ROOT / "geneconv_complete/dnm_analysis/export"
OUT_DIR  = PROJECT_ROOT / "geneconv_complete/__PAPER_3/TablesFinal"

AF_LABELS = {
    "0_1e-05":     "<10⁻⁵",
    "1e-05_00001": "10⁻⁵–10⁻⁴",
    "00001_0001":  "10⁻⁴–10⁻³",
    "0001_0005":   "10⁻³–5×10⁻³",
    "0005_001":    "5×10⁻³–10⁻²",
    "001_005":     "10⁻²–5×10⁻²",
    "005_01":      "5×10⁻²–0.1",
    "01_05":       "0.1–0.5",
    "05_2":        ">0.5",
}
AF_ORDER = list(AF_LABELS)
K_VALUES = [17, 19, 21, 31, 41, 51, 61, 71, 81, 91]
N_TESTS  = len(K_VALUES) * len(AF_ORDER)


def bonferroni_log10p(s):
    if s is None or s == "":
        return ""
    return round(min(float(s) + math.log10(N_TESTS), 0.0), 1)


def read_rout(filepath):
    rows = []
    with open(filepath) as f:
        header = f.readline().strip().split("\t")
        for line in f:
            vals = line.strip().split("\t")
            if len(vals) < len(header):
                continue
            r = dict(zip(header, vals))
            r["log_p_positional"] = vals[7]
            r["log_p_concordance"] = vals[-1]
            rows.append(r)
    return rows


def build_panel(rout_rows):
    panel = []
    for row in rout_rows:
        k = int(row["RepLength"])
        af = row["FrequencyInterval"]
        if k not in K_VALUES or af not in AF_LABELS:
            continue
        panel.append({
            "k": k, "af": af, "af_label": AF_LABELS[af],
            "a": int(row["NrOfVarPosConvPos"]),
            "b": int(row["NrOfNoVarPosConvPos"]),
            "c": int(row["NrOfVarPosNoConvPos"]),
            "d": int(row["NrOfNoVarPosNoConvPos"]),
            "log_p_pos": bonferroni_log10p(row["log_p_positional"]),
            "concor_var":   int(row["ConcorVar"]),
            "concor_novar": int(row["ConcorNoVar"]),
            "discor_var":   int(row["DiscorVar"]),
            "discor_novar": int(row["DiscorNoVar"]),
            "log_p_conc": bonferroni_log10p(row["log_p_concordance"]),
        })
    panel.sort(key=lambda x: (x["k"], AF_ORDER.index(x["af"])))
    return panel


def write_xlsx(data, st_num, title, outpath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Supplementary Table {st_num}"
    bold = Font(bold=True)
    header_font = Font(bold=True, size=11)
    thin_border = Border(bottom=Side(style="thin"))

    ws.append([f"Supplementary Table {st_num}: {title}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws["A1"].font = header_font
    ws.append([])

    row_num = ws.max_row + 1
    ws.cell(row=row_num, column=3, value="Positional enrichment (2×2 table)").font = bold
    ws.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=7)
    ws.cell(row=row_num, column=8, value="Concordance (2×2 table)").font = bold
    ws.merge_cells(start_row=row_num, start_column=8, end_row=row_num, end_column=12)

    ws.append([
        "k", "AF bin",
        "Var ∩ GC", "NoVar ∩ GC", "Var ∩ nonGC", "NoVar ∩ nonGC",
        "log₁₀(p) Bonf.",
        "Concor. var", "Concor. no var", "Discord. var", "Discord. no var",
        "log₁₀(p) Bonf.",
    ])
    for col in range(1, 13):
        ws.cell(row=ws.max_row, column=col).font = bold
        ws.cell(row=ws.max_row, column=col).border = thin_border

    for entry in data:
        ws.append([entry["k"], entry["af_label"],
                   entry["a"], entry["b"], entry["c"], entry["d"],
                   entry["log_p_pos"],
                   entry["concor_var"], entry["concor_novar"],
                   entry["discor_var"], entry["discor_novar"],
                   entry["log_p_conc"]])

    col_widths = [6, 16, 12, 16, 14, 18, 14, 12, 16, 12, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    outpath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outpath)
    print(f"Saved: {outpath}")


for st_num, src_subdir, title in [
    (7, "stats_allmapp",     "De novo mutation enrichment at gene conversion-compatible positions."),
    (8, "stats_nosegdupmapp", "De novo mutation enrichment outside segmental duplications."),
]:
    src = DNM_BASE / src_subdir / "_All_1000000_1000000_ALL_StatsSummary.R.out.txt"
    panel = build_panel(read_rout(src))
    print(f"ST{st_num}: {len(panel)} rows")
    write_xlsx(panel, st_num, title, OUT_DIR / f"SupplementaryTable_{st_num}.xlsx")
